"""
Report generation service for HAZOP studies.

Generates Excel worksheets and PDF reports from study data
using openpyxl and reportlab.
"""

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
    Image,
)


# Risk color mapping
RISK_COLORS_EXCEL = {
    "critical": "FF0000",  # Red
    "high": "FF8C00",      # Dark orange
    "medium": "FFD700",    # Gold/yellow
    "low": "32CD32",       # Lime green
}

RISK_COLORS_PDF = {
    "critical": colors.red,
    "high": colors.orange,
    "medium": colors.yellow,
    "low": colors.lightgreen,
}


def _get_risk_category(severity, likelihood) -> str:
    """Determine risk category from severity and likelihood."""
    likelihood_map = {"A": 5, "B": 4, "C": 3, "D": 2, "E": 1}
    sev = int(severity) if severity else 0
    lik = likelihood_map.get(str(likelihood).upper(), 3) if likelihood else 1
    risk_val = sev * lik

    if risk_val >= 15:
        return "critical"
    elif risk_val >= 10:
        return "high"
    elif risk_val >= 5:
        return "medium"
    else:
        return "low"


class ReportGenerator:
    """Generates Excel and PDF HAZOP study reports."""

    def generate_excel(self, study_data: dict) -> bytes:
        """
        Generate a comprehensive HAZOP Excel worksheet.

        Contents:
        - Cover sheet with study metadata
        - Node-by-node worksheets with all HAZOP columns
        - Summary sheet with risk distribution

        Args:
            study_data: Complete study data including nodes and deviations.

        Returns:
            Excel file as bytes.
        """
        wb = Workbook()

        # ---- Styles ----
        header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(
            start_color="2F5496", end_color="2F5496", fill_type="solid"
        )
        title_font = Font(name="Calibri", bold=True, size=18)
        subtitle_font = Font(name="Calibri", bold=True, size=14)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        wrap_alignment = Alignment(wrap_text=True, vertical="top")

        # ---- Cover Sheet ----
        ws_cover = wb.active
        ws_cover.title = "Cover"
        ws_cover.sheet_properties.tabColor = "2F5496"

        ws_cover.merge_cells("B2:F2")
        ws_cover["B2"] = "HAZOP STUDY WORKSHEET"
        ws_cover["B2"].font = title_font

        cover_fields = [
            ("Study Name", study_data.get("name", "Untitled Study")),
            ("Description", study_data.get("description", "")),
            ("Process Type", study_data.get("process_type", "")),
            ("Status", study_data.get("status", "")),
            ("Created", study_data.get("created_at", "")),
            ("Study ID", study_data.get("study_id", "")),
            (
                "Total Nodes",
                str(len(study_data.get("nodes", []))),
            ),
            (
                "Total Deviations",
                str(
                    sum(
                        len(n.get("deviations", []))
                        for n in study_data.get("nodes", [])
                    )
                ),
            ),
            (
                "Report Generated",
                datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
            ),
        ]

        for i, (label, value) in enumerate(cover_fields, start=4):
            ws_cover[f"B{i}"] = label
            ws_cover[f"B{i}"].font = Font(bold=True)
            ws_cover[f"C{i}"] = value

        ws_cover.column_dimensions["B"].width = 20
        ws_cover.column_dimensions["C"].width = 60

        # ---- Node Worksheets ----
        nodes = study_data.get("nodes", [])
        for node_idx, node in enumerate(nodes):
            node_name = node.get("name", f"Node {node_idx + 1}")
            # Excel sheet name max 31 chars
            sheet_name = node_name[:28] if len(node_name) > 28 else node_name
            # Sanitize sheet name
            for char in ["\\", "/", "*", "?", ":", "[", "]"]:
                sheet_name = sheet_name.replace(char, "_")

            ws = wb.create_sheet(title=sheet_name)
            ws.sheet_properties.tabColor = "4472C4"

            # Node header
            ws.merge_cells("A1:K1")
            ws["A1"] = f"HAZOP Worksheet - {node_name}"
            ws["A1"].font = subtitle_font

            ws["A2"] = f"Equipment Type: {node.get('equipment_type', 'N/A')}"
            ws["A3"] = (
                f"Operating Conditions: "
                f"{_format_conditions(node.get('operating_conditions', {}))}"
            )

            # Column headers
            headers = [
                "Guide Word",
                "Parameter",
                "Deviation",
                "Causes",
                "Consequences",
                "Existing Safeguards",
                "Severity",
                "Likelihood",
                "Risk Score",
                "Recommendations",
                "Notes",
            ]

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal="center", wrap_text=True
                )

            # Column widths
            col_widths = [12, 12, 18, 30, 30, 25, 10, 10, 10, 30, 20]
            for i, width in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Deviation rows
            deviations = node.get("deviations", [])
            for dev_idx, dev in enumerate(deviations):
                row = 6 + dev_idx

                severity = dev.get("severity", 0)
                likelihood = dev.get("likelihood", "")
                risk_cat = _get_risk_category(severity, likelihood)

                row_data = [
                    dev.get("guide_word", ""),
                    dev.get("parameter", ""),
                    f"{dev.get('guide_word', '')} {dev.get('parameter', '')}",
                    _safe_join(dev.get("causes")),
                    _safe_join(dev.get("consequences")),
                    _safe_join(dev.get("safeguards")),
                    severity,
                    likelihood,
                    dev.get("risk_score", ""),
                    _safe_join(dev.get("recommendations")),
                    str(dev.get("notes", "") or ""),
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = wrap_alignment

                    # Color code risk score cell
                    if col_idx == 9 and risk_cat in RISK_COLORS_EXCEL:
                        cell.fill = PatternFill(
                            start_color=RISK_COLORS_EXCEL[risk_cat],
                            end_color=RISK_COLORS_EXCEL[risk_cat],
                            fill_type="solid",
                        )

        # ---- Summary Sheet ----
        ws_summary = wb.create_sheet(title="Summary")
        ws_summary.sheet_properties.tabColor = "FF0000"

        ws_summary.merge_cells("A1:E1")
        ws_summary["A1"] = "HAZOP Study Risk Summary"
        ws_summary["A1"].font = subtitle_font

        # Risk distribution
        summary_headers = [
            "Risk Category",
            "Count",
            "Percentage",
        ]
        for col_idx, header in enumerate(summary_headers, start=1):
            cell = ws_summary.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        risk_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}
        total_deviations = 0

        for node in nodes:
            for dev in node.get("deviations", []):
                total_deviations += 1
                severity = dev.get("severity", 0)
                likelihood = dev.get("likelihood", "")
                cat = _get_risk_category(severity, likelihood)
                risk_counts[cat] = risk_counts.get(cat, 0) + 1

        for i, (category, count) in enumerate(risk_counts.items()):
            row = 4 + i
            pct = (
                f"{count / total_deviations * 100:.1f}%"
                if total_deviations > 0
                else "0%"
            )

            ws_summary.cell(row=row, column=1, value=category.upper())
            ws_summary.cell(row=row, column=2, value=count)
            ws_summary.cell(row=row, column=3, value=pct)

            for col_idx in range(1, 4):
                cell = ws_summary.cell(row=row, column=col_idx)
                cell.border = thin_border
                if category in RISK_COLORS_EXCEL:
                    cell.fill = PatternFill(
                        start_color=RISK_COLORS_EXCEL[category],
                        end_color=RISK_COLORS_EXCEL[category],
                        fill_type="solid",
                    )

        ws_summary.column_dimensions["A"].width = 18
        ws_summary.column_dimensions["B"].width = 12
        ws_summary.column_dimensions["C"].width = 14

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def generate_pdf(self, study_data: dict) -> bytes:
        """
        Generate a comprehensive HAZOP PDF report.

        Contents:
        - Executive summary
        - Methodology description
        - Findings ranked by risk
        - Risk matrix visualization
        - Action items

        Args:
            study_data: Complete study data including nodes and deviations.

        Returns:
            PDF file as bytes.
        """
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=20 * mm,
            leftMargin=20 * mm,
            topMargin=25 * mm,
            bottomMargin=25 * mm,
        )

        styles = getSampleStyleSheet()
        elements = []

        # Custom styles
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Title"],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
        )
        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading1"],
            fontSize=16,
            spaceBefore=20,
            spaceAfter=10,
            textColor=colors.HexColor("#2F5496"),
        )
        heading2_style = ParagraphStyle(
            "CustomHeading2",
            parent=styles["Heading2"],
            fontSize=13,
            spaceBefore=12,
            spaceAfter=8,
            textColor=colors.HexColor("#4472C4"),
        )
        body_style = ParagraphStyle(
            "CustomBody",
            parent=styles["Normal"],
            fontSize=10,
            spaceAfter=6,
            alignment=TA_JUSTIFY,
        )

        # ---- Title Page ----
        elements.append(Spacer(1, 80))
        elements.append(Paragraph("HAZOP Study Report", title_style))
        elements.append(Spacer(1, 20))
        elements.append(
            Paragraph(
                study_data.get("name", "Untitled Study"),
                ParagraphStyle(
                    "Subtitle",
                    parent=styles["Heading2"],
                    alignment=TA_CENTER,
                    textColor=colors.HexColor("#4472C4"),
                ),
            )
        )
        elements.append(Spacer(1, 40))

        # Study metadata table
        meta_data = [
            ["Study ID", study_data.get("study_id", "N/A")],
            ["Process Type", study_data.get("process_type", "N/A")],
            ["Status", study_data.get("status", "N/A")],
            ["Created", study_data.get("created_at", "N/A")],
            [
                "Report Generated",
                datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
            ],
        ]
        meta_table = Table(meta_data, colWidths=[120, 350])
        meta_table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        elements.append(meta_table)
        elements.append(PageBreak())

        # ---- Executive Summary ----
        elements.append(Paragraph("1. Executive Summary", heading_style))

        nodes = study_data.get("nodes", [])
        total_deviations = sum(
            len(n.get("deviations", [])) for n in nodes
        )
        risk_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}

        for node in nodes:
            for dev in node.get("deviations", []):
                cat = _get_risk_category(
                    dev.get("severity", 0), dev.get("likelihood", "")
                )
                risk_counts[cat] += 1

        summary_text = (
            f"This HAZOP study analyzed {len(nodes)} process nodes and "
            f"identified {total_deviations} deviations. "
            f"Of these, {risk_counts['critical']} were rated as CRITICAL risk, "
            f"{risk_counts['high']} as HIGH risk, "
            f"{risk_counts['medium']} as MEDIUM risk, and "
            f"{risk_counts['low']} as LOW risk."
        )
        elements.append(Paragraph(summary_text, body_style))

        if study_data.get("description"):
            elements.append(Spacer(1, 10))
            elements.append(
                Paragraph(
                    f"<b>Process Description:</b> {study_data['description']}",
                    body_style,
                )
            )

        # Risk summary table
        elements.append(Spacer(1, 15))
        risk_table_data = [
            ["Risk Category", "Count", "Percentage"],
        ]
        for cat in ["critical", "high", "medium", "low"]:
            pct = (
                f"{risk_counts[cat] / total_deviations * 100:.1f}%"
                if total_deviations > 0
                else "0%"
            )
            risk_table_data.append([cat.upper(), str(risk_counts[cat]), pct])

        risk_table = Table(risk_table_data, colWidths=[120, 80, 80])
        risk_table_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ]
        # Color code risk rows
        risk_row_colors = {
            1: RISK_COLORS_PDF["critical"],
            2: RISK_COLORS_PDF["high"],
            3: RISK_COLORS_PDF["medium"],
            4: RISK_COLORS_PDF["low"],
        }
        for row_idx, color in risk_row_colors.items():
            risk_table_style.append(
                ("BACKGROUND", (0, row_idx), (-1, row_idx), color)
            )

        risk_table.setStyle(TableStyle(risk_table_style))
        elements.append(risk_table)

        # ---- Methodology ----
        elements.append(PageBreak())
        elements.append(Paragraph("2. Methodology", heading_style))
        methodology_text = (
            "This study was conducted following the HAZOP methodology as "
            "defined in IEC 61882. Each process node was systematically "
            "analyzed by applying guide words (NO, MORE, LESS, AS WELL AS, "
            "PART OF, REVERSE, OTHER THAN) to process parameters (Flow, "
            "Temperature, Pressure, Level, Composition) to identify "
            "potential deviations from design intent. For each deviation, "
            "causes, consequences, existing safeguards, and recommendations "
            "were documented."
        )
        elements.append(Paragraph(methodology_text, body_style))

        # Risk matrix
        elements.append(Spacer(1, 15))
        elements.append(Paragraph("Risk Matrix", heading2_style))

        matrix_data = [
            ["", "A\nExtremely\nUnlikely", "B\nRemote", "C\nUnlikely", "D\nLikely", "E\nFrequent"],
            ["5 - Catastrophic", "5A", "5B", "5C", "5D", "5E"],
            ["4 - Major", "4A", "4B", "4C", "4D", "4E"],
            ["3 - Moderate", "3A", "3B", "3C", "3D", "3E"],
            ["2 - Minor", "2A", "2B", "2C", "2D", "2E"],
            ["1 - Negligible", "1A", "1B", "1C", "1D", "1E"],
        ]
        matrix_table = Table(
            matrix_data, colWidths=[90, 65, 65, 65, 65, 65]
        )
        matrix_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#2F5496")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]

        # Color the risk matrix cells
        risk_matrix_colors = {
            (1, 1): "low", (1, 2): "low", (1, 3): "low",
            (1, 4): "medium", (1, 5): "medium",
            (2, 1): "low", (2, 2): "low", (2, 3): "medium",
            (2, 4): "medium", (2, 5): "high",
            (3, 1): "low", (3, 2): "medium", (3, 3): "medium",
            (3, 4): "high", (3, 5): "high",
            (4, 1): "medium", (4, 2): "medium", (4, 3): "high",
            (4, 4): "high", (4, 5): "critical",
            (5, 1): "medium", (5, 2): "high", (5, 3): "high",
            (5, 4): "critical", (5, 5): "critical",
        }
        for (col, row), cat in risk_matrix_colors.items():
            # Table rows are inverted: severity 5 is row 1, severity 1 is row 5
            table_row = 6 - row
            matrix_style.append(
                (
                    "BACKGROUND",
                    (col, table_row),
                    (col, table_row),
                    RISK_COLORS_PDF[cat],
                )
            )

        matrix_table.setStyle(TableStyle(matrix_style))
        elements.append(matrix_table)

        # ---- Findings ----
        elements.append(PageBreak())
        elements.append(
            Paragraph("3. Findings by Risk Level", heading_style)
        )

        # Collect and sort all deviations by risk
        all_deviations = []
        for node in nodes:
            for dev in node.get("deviations", []):
                all_deviations.append(
                    {
                        **dev,
                        "node_name": node.get("name", "Unknown"),
                        "equipment_type": node.get("equipment_type", ""),
                    }
                )

        risk_priority = {"critical": 0, "high": 1, "medium": 2, "low": 3}
        all_deviations.sort(
            key=lambda d: risk_priority.get(
                _get_risk_category(
                    d.get("severity", 0), d.get("likelihood", "")
                ),
                4,
            )
        )

        for dev in all_deviations:
            cat = _get_risk_category(
                dev.get("severity", 0), dev.get("likelihood", "")
            )
            elements.append(
                Paragraph(
                    f"<b>[{cat.upper()}]</b> {_escape_xml(dev.get('node_name', ''))} - "
                    f"{_escape_xml(dev.get('guide_word', ''))} {_escape_xml(dev.get('parameter', ''))}",
                    heading2_style,
                )
            )

            causes = _safe_list(dev.get("causes"))
            if causes:
                elements.append(
                    Paragraph(
                        "<b>Causes:</b> " + _escape_xml("; ".join(causes)),
                        body_style,
                    )
                )
            consequences = _safe_list(dev.get("consequences"))
            if consequences:
                elements.append(
                    Paragraph(
                        "<b>Consequences:</b> " + _escape_xml("; ".join(consequences)),
                        body_style,
                    )
                )
            safeguards = _safe_list(dev.get("safeguards"))
            if safeguards:
                elements.append(
                    Paragraph(
                        "<b>Safeguards:</b> " + _escape_xml("; ".join(safeguards)),
                        body_style,
                    )
                )
            recommendations = _safe_list(dev.get("recommendations"))
            if recommendations:
                elements.append(
                    Paragraph(
                        "<b>Recommendations:</b> " + _escape_xml("; ".join(recommendations)),
                        body_style,
                    )
                )

            elements.append(
                Paragraph(
                    f"<b>Risk Score:</b> {dev.get('risk_score', 'N/A')} "
                    f"(Severity: {dev.get('severity', 'N/A')}, "
                    f"Likelihood: {dev.get('likelihood', 'N/A')})",
                    body_style,
                )
            )
            elements.append(Spacer(1, 8))

        # ---- Action Items ----
        elements.append(PageBreak())
        elements.append(Paragraph("4. Action Items", heading_style))

        action_items = []
        for dev in all_deviations:
            cat = _get_risk_category(
                dev.get("severity", 0), dev.get("likelihood", "")
            )
            for rec in _safe_list(dev.get("recommendations")):
                action_items.append(
                    {
                        "priority": cat,
                        "node": dev.get("node_name", ""),
                        "deviation": (
                            f"{dev.get('guide_word', '')} "
                            f"{dev.get('parameter', '')}"
                        ),
                        "action": rec,
                    }
                )

        if action_items:
            action_table_data = [
                ["Priority", "Node", "Deviation", "Action Required"],
            ]
            for item in action_items:
                action_table_data.append(
                    [
                        item["priority"].upper(),
                        Paragraph(_escape_xml(item["node"]), body_style),
                        Paragraph(_escape_xml(item["deviation"]), body_style),
                        Paragraph(_escape_xml(item["action"]), body_style),
                    ]
                )

            action_table = Table(
                action_table_data,
                colWidths=[60, 100, 80, 270],
            )
            action_style = [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#2F5496"),
                ),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
            ]
            action_table.setStyle(TableStyle(action_style))
            elements.append(action_table)
        else:
            elements.append(
                Paragraph("No action items identified.", body_style)
            )

        # Build PDF
        doc.build(elements)
        return output.getvalue()


def _format_conditions(conditions) -> str:
    """Format operating conditions as a readable string."""
    if not conditions:
        return "N/A"
    if isinstance(conditions, str):
        return conditions
    if isinstance(conditions, dict):
        parts = [f"{k}: {v}" for k, v in conditions.items() if v]
        return ", ".join(parts) if parts else "N/A"
    return str(conditions)


def _safe_list(val) -> list:
    """Ensure value is a list of strings."""
    if not val:
        return []
    if isinstance(val, str):
        return [val]
    if isinstance(val, list):
        return [str(item) for item in val if item]
    return [str(val)]


def _safe_join(val, separator="\n") -> str:
    """Safely join a value that might be a list, string, or other type."""
    items = _safe_list(val)
    return separator.join(items) if items else ""


def _escape_xml(text: str) -> str:
    """Escape special XML characters for reportlab Paragraph."""
    if not text:
        return ""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


# Module-level singleton
report_generator = ReportGenerator()
